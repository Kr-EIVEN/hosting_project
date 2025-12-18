""" 주제3_back_data / 주제3_결산보고서양식 기반 자동 결산보고서 생성 스크립트
(수식 자동 파싱 + direct 셀 참조 + C열 참조 수식 처리 버전, 조건별 컬럼까지 한 파일에 통합 저장)
"""

import os
import re
import pandas as pd
import openpyxl

# -------------------------------------------------------------------
# 1. 파일 경로 (flaskbackend 기준)
# -------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # .../flaskbackend
DATA_DIR = os.path.join(BASE_DIR, "report_data")

BACK_DATA_FILE = os.path.join(DATA_DIR, "주제3_back_data.xlsx")
TEMPLATE_FILE  = os.path.join(DATA_DIR, "주제3_결산보고서양식.xlsx")
OUTPUT_FILE    = os.path.join(DATA_DIR, "주제3_결산보고서_통합.xlsx")

# -------------------------------------------------------------------
# 2. 분류 기준 (기존 7개 + 전체 + 새 2개)
# -------------------------------------------------------------------
GROUPING_CONFIG = {
    "전체": None,          # 전체 합산용
    "플랜트": "플랜트",
    "대표차종": "대표차종",
    "유통경로": "유통 경로",
    "판매문서유형": "판매 문서 유형",
    "기타매출유형": "기타매출유형",
    "레코드유형": "레코드 유형",
    "평가클래스": "평가클래스",
    # [OK] 추가 조건 2개
    "Prod.계층구조01-2": "Prod.계층구조01-2",
    "손익센터": "손익 센터",
}

# -------------------------------------------------------------------
# 3. 엑셀 열 문자 → 인덱스 / 컬럼명 변환 유틸
# -------------------------------------------------------------------
def excel_col_to_index(col: str) -> int:
    col = col.upper()
    idx = 0
    for c in col:
        if not ('A' <= c <= 'Z'):
            continue
        idx = idx * 26 + (ord(c) - ord('A') + 1)
    return idx


def build_backdata_colmap(back_data_path: str, sheet_name: int | str = 0) -> dict:
    wb = openpyxl.load_workbook(back_data_path, data_only=False)
    ws = wb[wb.sheetnames[sheet_name]]

    colmap: dict[str, str] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        col_letter = cell.column_letter
        col_name = str(cell.value).strip()
        colmap[col_letter] = col_name

    return colmap


# -------------------------------------------------------------------
# 4. SUMIFS 수식 파싱 유틸
# -------------------------------------------------------------------
def split_top_level_terms(expr: str):
    expr = expr.replace(" ", "")
    terms = []
    buf = []
    depth = 0
    sign = 1

    i = 0
    while i < len(expr):
        ch = expr[i]
        if ch == '(':
            depth += 1
            buf.append(ch)
        elif ch == ')':
            depth -= 1
            buf.append(ch)
        elif depth == 0 and ch in ['+', '-']:
            if buf:
                term_str = "".join(buf)
                if term_str:
                    terms.append((sign, term_str))
                buf = []
            sign = 1 if ch == '+' else -1
        else:
            buf.append(ch)
        i += 1

    if buf:
        term_str = "".join(buf)
        if term_str:
            terms.append((sign, term_str))

    return terms


def parse_sumifs_term(term: str, colmap: dict[str, str]):
    if not term.upper().startswith("SUMIFS("):
        return None

    inner = term[7:-1]

    m_range = re.search(r"\$([A-Z]{1,2})\$2:\$[A-Z]{1,2}\$1073", inner)
    if not m_range:
        return None
    value_col_letter = m_range.group(1)
    value_col_name = colmap.get(value_col_letter)
    if value_col_name is None:
        return None

    crit_pattern = re.compile(
        r"\$([A-Z]{1,2})\$2:\$[A-Z]{1,2}\$1073,\"([^\"]*)\""
    )

    criteria: dict[str, list[str]] = {}
    for col_letter, crit_val in crit_pattern.findall(inner):
        col_name = colmap.get(col_letter)
        if col_name is None:
            continue
        criteria.setdefault(col_name, []).append(crit_val)

    return {
        "value_col": value_col_name,
        "criteria": criteria,
    }


def parse_sumifs_formula(
    formula: str,
    colmap: dict[str, str],
):
    if not isinstance(formula, str):
        return []

    if formula.startswith("="):
        expr = formula[1:]
    else:
        expr = formula

    terms = split_top_level_terms(expr)
    parsed_terms = []

    for sign, term_str in terms:
        parsed = parse_sumifs_term(term_str, colmap)
        if not parsed:
            continue
        parsed_terms.append(
            {
                "sign": sign,
                "value_col": parsed["value_col"],
                "criteria": parsed["criteria"],
            }
        )

    return parsed_terms


# -------------------------------------------------------------------
# 5. Back Data에서 파싱된 SUMIFS 구조를 실제 값으로 평가
# -------------------------------------------------------------------
def apply_criteria(df: pd.DataFrame, col: str, values: list[str]) -> pd.Series:
    if col not in df.columns or not values:
        return pd.Series(False, index=df.index)

    ser = df[col]
    ser_str = ser.astype(str).str.rstrip("0").str.rstrip(".")

    mask = pd.Series(False, index=df.index)
    for v in values:
        v = str(v)
        mask |= (ser_str == v)

    return mask


def eval_sumifs_terms(
    df: pd.DataFrame,
    terms: list[dict],
) -> float:
    total = 0.0

    for term in terms:
        sign = term["sign"]
        value_col = term["value_col"]
        criteria = term["criteria"]

        if value_col not in df.columns:
            continue

        if not criteria:
            mask = pd.Series(True, index=df.index)
        else:
            masks = []
            for col_name, vals in criteria.items():
                m = apply_criteria(df, col_name, vals)
                masks.append(m)
            mask = masks[0]
            for m in masks[1:]:
                mask &= m

        subtotal = df.loc[mask, value_col].sum()
        total += sign * float(subtotal)

    return float(total)


# -------------------------------------------------------------------
# 5-1. Back Data 특정 셀 직접 참조 수식 처리
# -------------------------------------------------------------------
def eval_direct_backdata_cell(
    formula: str,
    ws_back: openpyxl.worksheet.worksheet.Worksheet,
) -> float | None:
    if not isinstance(formula, str):
        return None
    m = re.search(r"!\$[OK]([A-Z]{1,3})\$[OK](\d+)", formula)
    if not m:
        return None

    col_letter = m.group(1)
    row_num = int(m.group(2))
    cell_ref = f"{col_letter}{row_num}"

    cell = ws_back[cell_ref]
    val = cell.value
    if val is None:
        return 0.0
    try:
        return float(val)
    except Exception:
        return 0.0


# -------------------------------------------------------------------
# 5-2. 템플릿 C열 내 내부 셀 참조 수식 → 번호 기반 산출식
# -------------------------------------------------------------------
def convert_c_formula_to_num_expr(
    formula: str,
    excel_row_to_num: dict[int, int],
) -> str | None:
    if not isinstance(formula, str):
        return None
    if not formula.startswith("="):
        return None

    expr = formula[1:]

    sum_pattern = re.compile(
        r"SUM\(\s*\$[OK]C\$[OK](\d+)\s*:\s*\$[OK]C\$[OK](\d+)\s*\)",
        flags=re.I,
    )

    def sum_repl(m: re.Match) -> str:
        row1 = int(m.group(1))
        row2 = int(m.group(2))
        n1 = excel_row_to_num.get(row1)
        n2 = excel_row_to_num.get(row2)
        if n1 is None or n2 is None:
            return "0"
        return f"{n1}+…+{n2}"

    expr2, n_sub_sum = sum_pattern.subn(sum_repl, expr)

    cref_pattern = re.compile(r"\$[OK]C\$[OK](\d+)", flags=re.I)

    def c_repl(m: re.Match) -> str:
        row = int(m.group(1))
        n = excel_row_to_num.get(row)
        if n is None:
            return "0"
        return str(n)

    expr3, n_sub_c = cref_pattern.subn(c_repl, expr2)

    if n_sub_sum == 0 and n_sub_c == 0:
        return None

    return expr3


# -------------------------------------------------------------------
# 6. 산출식(번호 조합) 평가
# -------------------------------------------------------------------
def eval_formula(expr: str, values: dict[int, float]) -> float:
    if not isinstance(expr, str) or not expr:
        return 0.0

    expr = expr.replace(" ", "")

    m = re.fullmatch(r"(\d+)\+…\+(\d+)", expr)
    if m:
        start = int(m.group(1))
        end = int(m.group(2))
        return sum(values.get(i, 0.0) for i in range(start, end + 1))

    def repl(match: re.Match) -> str:
        n = int(match.group(0))
        return f"values.get({n}, 0.0)"

    python_expr = re.sub(r"\d+", repl, expr)

    return float(eval(python_expr, {"__builtins__": {}}, {"values": values}))


# -------------------------------------------------------------------
# 7. 한 그룹에 대한 "번호 → 금액" 계산
# -------------------------------------------------------------------
def compute_values_for_group(
    back_df: pd.DataFrame,
    tpl_rows: pd.DataFrame,
    ws_tpl: openpyxl.worksheet.worksheet.Worksheet,
    colmap_back: dict[str, str],
    group_col: str | None,
    group_value,
    ws_back_data: openpyxl.worksheet.worksheet.Worksheet,
    excel_row_to_num: dict[int, int],
) -> dict[int, float]:
    if group_col is None:
        df_group = back_df.copy()
    else:
        df_group = back_df[back_df[group_col] == group_value].copy()

    values: dict[int, float] = {}

    # 판매수량(국내/수출) SUMIFS 셋업
    qty_terms_domestic = None
    qty_terms_export = None
    try:
        dom_formula = (
            "=SUMIFS('Back data'!$S$2:$S$1073,'Back data'!$O$2:$O$1073,\"F\","
            "'Back data'!$K$2:$K$1073,\"10\")"
            "+SUMIFS('Back data'!$S$2:$S$1073,'Back data'!$O$2:$O$1073,\"F\","
            "'Back data'!$K$2:$K$1073,\"20\")"
            "+SUMIFS('Back data'!$S$2:$S$1073,'Back data'!$O$2:$O$1073,\"F\","
            "'Back data'!$K$2:$K$1073,\"91\")"
            "-SUMIFS('Back data'!$S$2:$S$1073,'Back data'!$O$2:$O$1073,\"F\","
            "'Back data'!$K$2:$K$1073,\"10\",'Back data'!$I$2:$I$1073,\"3100\")"
            "-SUMIFS('Back data'!$S$2:$S$1073,'Back data'!$O$2:$O$1073,\"F\","
            "'Back data'!$K$2:$K$1073,\"20\",'Back data'!$I$2:$I$1073,\"3100\")"
            "-SUMIFS('Back data'!$S$2:$S$1073,'Back data'!$O$2:$O$1073,\"F\","
            "'Back data'!$K$2:$K$1073,\"91\",'Back data'!$I$2:$I$1073,\"3100\")"
        )
        exp_formula = (
            "=SUMIFS('Back data'!$S$2:$S$1073,'Back data'!$O$2:$O$1073,\"F\","
            "'Back data'!$K$2:$K$1073,\"30\")"
            "-SUMIFS('Back data'!$S$2:$S$1073,'Back data'!$O$2:$O$1073,\"F\","
            "'Back data'!$K$2:$K$1073,\"30\",'Back data'!$I$2:$I$1073,\"3100\")"
        )

        dom_terms_raw = parse_sumifs_formula(dom_formula, colmap_back)
        if dom_terms_raw:
            qty_terms_domestic = [
                {
                    "sign": t["sign"],
                    "value_col": "판매수량",
                    "criteria": t["criteria"],
                }
                for t in dom_terms_raw
            ]

        exp_terms_raw = parse_sumifs_formula(exp_formula, colmap_back)
        if exp_terms_raw:
            qty_terms_export = [
                {
                    "sign": t["sign"],
                    "value_col": "판매수량",
                    "criteria": t["criteria"],
                }
                for t in exp_terms_raw
            ]
    except Exception:
        qty_terms_domestic = None
        qty_terms_export = None

    # 1차: 기본 항목 계산
    for _, row in tpl_rows.iterrows():
        num = int(row["번호"])
        excel_row = num + 2
        formula = ws_tpl.cell(excel_row, 3).value
        item_name = str(row.get("항목"))

        val = None

        if item_name == "판매수량(국내)" and qty_terms_domestic is not None:
            val = eval_sumifs_terms(df_group, qty_terms_domestic)
        elif item_name == "판매수량(수출)" and qty_terms_export is not None:
            val = eval_sumifs_terms(df_group, qty_terms_export)
        else:
            parsed_terms = parse_sumifs_formula(formula, colmap_back) if isinstance(formula, str) else []
            if parsed_terms:
                val = eval_sumifs_terms(df_group, parsed_terms)
            else:
                if group_col is None:
                    direct_val = eval_direct_backdata_cell(formula, ws_back_data)
                    if direct_val is not None:
                        val = direct_val
                    else:
                        value_field_name = row.get("값필드명")
                        if pd.notna(value_field_name):
                            value_field_name = str(value_field_name)
                            if value_field_name in df_group.columns:
                                val = float(df_group[value_field_name].sum())
                            else:
                                val = 0.0
                else:
                    value_field_name = row.get("값필드명")
                    if pd.notna(value_field_name):
                        value_field_name = str(value_field_name)
                        if value_field_name in df_group.columns:
                            val = float(df_group[value_field_name].sum())
                        else:
                            val = 0.0

        if val is not None:
            values[num] = val

    # 2차: 산출식
    remaining: dict[int, str] = {}

    for _, row in tpl_rows.iterrows():
        num = int(row["번호"])
        excel_row = num + 2
        expr: str | None = None

        if pd.notna(row.get("산출식")):
            expr = str(row["산출식"])
        else:
            formula = ws_tpl.cell(excel_row, 3).value
            expr_from_c = convert_c_formula_to_num_expr(formula, excel_row_to_num)
            if expr_from_c:
                expr = expr_from_c

        if expr:
            remaining[num] = expr

    def can_eval_expr(expr: str,
                      values_dict: dict[int, float],
                      remaining_keys: set[int]) -> bool:
        nums = set(int(m.group(0)) for m in re.finditer(r"\d+", expr))
        if not nums:
            return False

        for n in nums:
            if n in remaining_keys:
                return False
            if n not in values_dict:
                return False

        return True

    max_iter = len(remaining) + 5
    for _ in range(max_iter):
        if not remaining:
            break

        done_this_round: list[int] = []

        keys_snapshot = set(remaining.keys())

        for num, expr in list(remaining.items()):
            if not can_eval_expr(expr, values, keys_snapshot):
                continue

            try:
                result = eval_formula(expr, values)
            except Exception:
                continue
            else:
                values[num] = result
                done_this_round.append(num)

        if not done_this_round:
            break

        for num in done_this_round:
            remaining.pop(num, None)

    for num in remaining.keys():
        values.setdefault(num, 0.0)

    try:
        item_to_num = {
            str(r["항목"]): int(r["번호"])
            for _, r in tpl_rows.iterrows()
            if pd.notna(r.get("항목")) and pd.notna(r.get("번호"))
        }
        n_pre_tax = item_to_num.get("법인세차감전순이익")
        n_tax = item_to_num.get("법인세비용")
        n_ni = item_to_num.get("당기순이익")

        if n_pre_tax is not None and n_tax is not None and n_ni is not None:
            pre_tax_val = values.get(n_pre_tax, 0.0)
            tax_val = values.get(n_tax, 0.0)
            values[n_ni] = pre_tax_val - tax_val
    except Exception:
        pass

    return values


# -------------------------------------------------------------------
# 8. 전체 / 분류조건별 컬럼을 한 DF로 통합
# -------------------------------------------------------------------
def generate_pl_report_df(
    back_data_file: str = BACK_DATA_FILE,
    template_file: str = TEMPLATE_FILE,
) -> pd.DataFrame:
    back_df = pd.read_excel(back_data_file, sheet_name=0)
    tpl_df  = pd.read_excel(template_file, sheet_name=0)

    wb_tpl = openpyxl.load_workbook(template_file, data_only=False)
    ws_tpl = wb_tpl[wb_tpl.sheetnames[0]]

    wb_back = openpyxl.load_workbook(back_data_file, data_only=True)
    ws_back = wb_back[wb_back.sheetnames[0]]

    colmap_back = build_backdata_colmap(back_data_file, sheet_name=0)

    tpl_rows = tpl_df[tpl_df["번호"].notna()].copy()

    num_to_item = {
        int(row["번호"]): str(row["항목"]) for _, row in tpl_rows.iterrows()
    }
    ordered_numbers = [int(n) for n in tpl_rows["번호"].tolist()]

    excel_row_to_num: dict[int, int] = {}
    for n in ordered_numbers:
        excel_row = n + 2
        excel_row_to_num[excel_row] = n

    integrated_df = pd.DataFrame(
        {
            "번호": ordered_numbers,
            "항목": [num_to_item[n] for n in ordered_numbers],
        }
    )

    for cond_name, group_col in GROUPING_CONFIG.items():
        if group_col is None:
            group_values = [None]
            group_labels = ["전체"]
        else:
            uniques = (
                back_df[group_col]
                .dropna()
                .drop_duplicates()
                .tolist()
            )
            group_values = sorted(uniques)
            group_labels = []
            for v in group_values:
                if isinstance(v, (int, float)) and pd.notna(v) and float(v).is_integer():
                    group_labels.append(str(int(v)))
                else:
                    group_labels.append(str(v))

        cond_cols: list[str] = []

        for gv, label in zip(group_values, group_labels):
            values_dict = compute_values_for_group(
                back_df=back_df,
                tpl_rows=tpl_rows,
                ws_tpl=ws_tpl,
                colmap_back=colmap_back,
                group_col=group_col,
                group_value=gv,
                ws_back_data=ws_back,
                excel_row_to_num=excel_row_to_num,
            )

            col_values = [values_dict.get(n, 0.0) for n in ordered_numbers]

            if cond_name == "전체":
                col_name = label
            else:
                col_name = f"{cond_name}_{label}"

            integrated_df[col_name] = col_values

            if cond_name != "전체":
                cond_cols.append(col_name)

        if cond_name != "전체" and cond_cols:
            total_col_name = f"{cond_name}_전체"
            integrated_df[total_col_name] = integrated_df[cond_cols].sum(axis=1)

        print(f"[완료] 조건 '{cond_name}' 컬럼 병합 완료")

    # [DOWN][DOWN][DOWN] 여기서부터 추가된 블록: 모든 금액/수량 컬럼 소수점 절사 [DOWN][DOWN][DOWN]
    for col in integrated_df.columns:
        if col in ["번호", "항목"]:
            continue
        if pd.api.types.is_numeric_dtype(integrated_df[col]):
            # NaN -> 0 처리 후 float → int 변환 (소수점 이하 버림, 0원 단위)
            integrated_df[col] = (
                integrated_df[col]
                .fillna(0)
                .astype(float)
                .astype("int64")
            )
    # [UP][UP][UP] 추가 끝 [UP][UP][UP]

    return integrated_df


def main():
    integrated_df = generate_pl_report_df()
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        integrated_df.to_excel(writer, sheet_name="보고서", index=False)
    print(f"[완료] '{OUTPUT_FILE}' 생성")


if __name__ == "__main__":
    main()
